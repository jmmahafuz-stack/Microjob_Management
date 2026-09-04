"""
Report Generation Service for exporting data in CSV, Excel, and PDF formats.
Generates payment, worker, job, commission, and financial reports.
"""

import io
import csv
from datetime import datetime, timedelta
from decimal import Decimal
from typing import List, Dict, Optional

from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

try:
    import reportlab
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
except ImportError:
    reportlab = None


class ReportGenerator:
    """Base class for generating reports in multiple formats."""
    
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S"
    CURRENCY = "৳"
    
    def __init__(self, title: str, start_date: Optional[datetime] = None, end_date: Optional[datetime] = None):
        """
        Initialize report generator.
        
        Args:
            title: Report title
            start_date: Filter start date (default: 30 days ago)
            end_date: Filter end date (default: today)
        """
        self.title = title
        self.end_date = end_date or timezone.now()
        self.start_date = start_date or (self.end_date - timedelta(days=30))
    
    def generate_csv_response(self, rows: List[List], headers: List[str], filename: str) -> HttpResponse:
        """Generate CSV response."""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        
        return response
    
    def generate_excel_response(self, rows: List[List], headers: List[str], filename: str) -> HttpResponse:
        """Generate Excel (.xlsx) response."""
        if not openpyxl:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Add title
        ws['A1'] = self.title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:J1')
        
        # Add date range
        date_str = f"Period: {self.start_date.date()} to {self.end_date.date()}"
        ws['A2'] = date_str
        ws['A2'].font = Font(italic=True, size=10)
        
        # Add headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data rows
        for row_num, row_data in enumerate(rows, 5):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Format numbers
                if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
                    if col_num > 2:  # Amount columns
                        cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            column_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = min(max_length + 5, 50)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=4, max_row=len(rows) + 4, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def generate_pdf_response(self, rows: List[List], headers: List[str], filename: str) -> HttpResponse:
        """Generate PDF response."""
        if not reportlab:
            raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=6,
        )
        elements.append(Paragraph(self.title, title_style))
        
        # Date range
        date_str = f"Period: {self.start_date.date()} to {self.end_date.date()}"
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=12,
        )
        elements.append(Paragraph(date_str, date_style))
        
        # Prepare table data
        table_data = [headers] + rows
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Generated timestamp
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        footer = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
        )
        elements.append(Paragraph(f"Generated on {ts}", footer))
        
        # Generate PDF
        doc.build(elements)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        
        return response

    def generate_sections_pdf_response(self, sections: List[Dict], filename: str) -> HttpResponse:
        """Generate one PDF containing multiple report sections."""
        if not reportlab:
            raise ImportError("reportlab is required for PDF export. Install it with: pip install reportlab")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            topMargin=0.45 * inch,
            bottomMargin=0.45 * inch,
            leftMargin=0.45 * inch,
            rightMargin=0.45 * inch,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(self.title, ParagraphStyle(
                'CombinedTitle', parent=styles['Heading1'], fontSize=18,
                textColor=colors.HexColor('#12384b'), spaceAfter=6,
            )),
            Paragraph(
                f"Period: {self.start_date.date()} to {self.end_date.date()}",
                ParagraphStyle('CombinedDate', parent=styles['Normal'], fontSize=9,
                               textColor=colors.grey, spaceAfter=14),
            ),
        ]

        for section in sections:
            elements.append(Paragraph(
                section['title'],
                ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=12,
                               textColor=colors.HexColor('#087f8c'), spaceBefore=8, spaceAfter=6),
            ))
            table_data = [section['headers']] + [
                [str(value) for value in row] for row in section['rows']
            ]
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#087f8c')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7.5),
                ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#cbd5e1')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f7f8')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.extend([table, Spacer(1, 0.16 * inch)])

        elements.append(Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            ParagraphStyle('CombinedFooter', parent=styles['Normal'], fontSize=8, textColor=colors.grey),
        ))
        doc.build(elements)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        return response


class PaymentReportGenerator(ReportGenerator):
    """Generate payment transaction reports."""
    
    def generate(self, format_type: str = 'excel') -> HttpResponse:
        """
        Generate payment report.
        
        Args:
            format_type: 'csv', 'excel', or 'pdf'
        
        Returns:
            HTTP response with file attachment
        """
        from payments.models import Payment
        
        # Get data
        payments = Payment.objects.filter(
            payment_date__range=[self.start_date, self.end_date]
        ).select_related('job__customer', 'job__worker').order_by('-payment_date')
        
        headers = [
            'Transaction ID', 'Date', 'Method', 'Customer Amount', 'Commission',
            'Worker Amount', 'Status', 'Worker Payout Status', 'Job Title', 'Customer', 'Worker'
        ]
        
        rows = []
        for payment in payments:
            rows.append([
                payment.transaction_id or 'N/A',
                payment.payment_date.strftime(self.DATE_FORMAT),
                payment.payment_method,
                f"{self.CURRENCY}{payment.customer_amount}",
                f"{self.CURRENCY}{payment.platform_commission}",
                f"{self.CURRENCY}{payment.worker_amount}",
                payment.payment_status,
                payment.worker_payout_status,
                payment.job.title if payment.job else 'N/A',
                payment.job.customer.get_full_name() if payment.job else 'N/A',
                payment.job.worker.get_full_name() if payment.job else 'N/A',
            ])
        
        filename = f"Payment_Report_{self.end_date.strftime('%Y%m%d')}.{format_type}"
        
        if format_type == 'csv':
            return self.generate_csv_response(rows, headers, filename)
        elif format_type == 'excel':
            return self.generate_excel_response(rows, headers, filename)
        elif format_type == 'pdf':
            return self.generate_pdf_response(rows, headers, filename)
        else:
            raise ValueError(f"Unsupported format: {format_type}")


class WorkerReportGenerator(ReportGenerator):
    """Generate worker earnings and performance reports."""
    
    def generate(self, format_type: str = 'excel') -> HttpResponse:
        """Generate worker report."""
        from workers.models import WorkerProfile
        from accounts.models import CustomUser
        
        workers = CustomUser.objects.filter(
            role='worker'
        ).select_related('worker_profile').annotate(
            jobs_completed=Count('jobs_completed', filter=Q(jobs_completed__status='COMPLETED')),
            avg_rating=Avg('reviews_received__rating')
        ).order_by('-worker_profile__total_earnings')
        
        headers = [
            'Name', 'Email', 'Verification Status', 'Jobs Completed', 'Average Rating',
            'Pending Earnings', 'Available Earnings', 'Withdrawn Earnings', 'Total Earnings',
            'Payment Method', 'Phone Number'
        ]
        
        rows = []
        for worker in workers:
            profile = worker.worker_profile
            phone = 'N/A'
            if profile.payout_method == 'BKash':
                phone = profile.bkash_number or 'N/A'
            elif profile.payout_method == 'Nagad':
                phone = profile.nagad_number or 'N/A'
            elif profile.payout_method == 'Rocket':
                phone = profile.rocket_number or 'N/A'
            
            rows.append([
                worker.get_full_name(),
                worker.email,
                profile.verification_status,
                worker.jobs_completed or 0,
                f"{profile.average_rating_cached:.1f}" if profile.average_rating_cached else 'N/A',
                f"{self.CURRENCY}{profile.pending_earnings}",
                f"{self.CURRENCY}{profile.available_earnings}",
                f"{self.CURRENCY}{profile.withdrawn_earnings}",
                f"{self.CURRENCY}{profile.total_earnings}",
                profile.payout_method,
                phone,
            ])
        
        filename = f"Worker_Report_{self.end_date.strftime('%Y%m%d')}.{format_type}"
        
        if format_type == 'csv':
            return self.generate_csv_response(rows, headers, filename)
        elif format_type == 'excel':
            return self.generate_excel_response(rows, headers, filename)
        else:
            raise ValueError(f"Format {format_type} not supported for worker reports")


class JobReportGenerator(ReportGenerator):
    """Generate job status and completion reports."""
    
    def generate(self, format_type: str = 'excel') -> HttpResponse:
        """Generate job report."""
        from bookings.models import Job
        
        jobs = Job.objects.filter(
            created_at__range=[self.start_date, self.end_date]
        ).select_related('customer', 'worker', 'service_request').order_by('-created_at')
        
        headers = [
            'Job ID', 'Title', 'Status', 'Customer', 'Worker', 'Price',
            'Created Date', 'Scheduled Date', 'Completed Date', 'Rating'
        ]
        
        rows = []
        for job in jobs:
            avg_rating = 'N/A'
            if job.worker:
                rating = job.worker.reviews_received.aggregate(Avg('rating'))['rating__avg']
                avg_rating = f"{rating:.1f}" if rating else 'N/A'
            
            rows.append([
                job.pk,
                job.title,
                job.status,
                job.customer.get_full_name(),
                job.worker.get_full_name() if job.worker else 'Unassigned',
                f"{self.CURRENCY}{job.proposed_price}",
                job.created_at.strftime(self.DATE_FORMAT),
                job.scheduled_date.strftime('%Y-%m-%d') if job.scheduled_date else 'N/A',
                job.completed_date.strftime(self.DATE_FORMAT) if hasattr(job, 'completed_date') else 'N/A',
                avg_rating,
            ])
        
        filename = f"Job_Report_{self.end_date.strftime('%Y%m%d')}.{format_type}"
        
        if format_type == 'csv':
            return self.generate_csv_response(rows, headers, filename)
        elif format_type == 'excel':
            return self.generate_excel_response(rows, headers, filename)
        else:
            raise ValueError(f"Format {format_type} not supported for job reports")


class FinancialReportGenerator(ReportGenerator):
    """Generate financial and commission reports."""
    
    def generate_commission_report(self, format_type: str = 'excel') -> HttpResponse:
        """Generate commission breakdown report."""
        from payments.models import Payment
        
        payments = Payment.objects.filter(
            payment_date__range=[self.start_date, self.end_date],
            payment_status='Verified'
        ).order_by('-payment_date')
        
        headers = ['Date', 'Transaction ID', 'Customer Amount', 'Commission Rate', 'Commission', 'Worker Amount']
        
        rows = []
        for payment in payments:
            rows.append([
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.transaction_id or 'N/A',
                f"{self.CURRENCY}{payment.customer_amount}",
                f"{payment.commission_rate}%",
                f"{self.CURRENCY}{payment.platform_commission}",
                f"{self.CURRENCY}{payment.worker_amount}",
            ])
        
        # Add totals
        totals = payments.aggregate(
            total_customer=Sum('customer_amount'),
            total_commission=Sum('platform_commission'),
            total_worker=Sum('worker_amount')
        )
        
        rows.append(['', 'TOTAL', 
            f"{self.CURRENCY}{totals['total_customer'] or 0}",
            '',
            f"{self.CURRENCY}{totals['total_commission'] or 0}",
            f"{self.CURRENCY}{totals['total_worker'] or 0}"])
        
        filename = f"Commission_Report_{self.end_date.strftime('%Y%m%d')}.{format_type}"
        
        if format_type == 'csv':
            return self.generate_csv_response(rows, headers, filename)
        elif format_type == 'excel':
            return self.generate_excel_response(rows, headers, filename)
        else:
            raise ValueError(f"Format {format_type} not supported")


class AllReportsGenerator(ReportGenerator):
    """Generate one PDF containing the platform's main report sections."""

    def generate(self, format_type: str = 'pdf') -> HttpResponse:
        if format_type != 'pdf':
            raise ValueError('The combined report is available as PDF only.')

        from accounts.models import CustomUser
        from bookings.models import Job
        from payments.models import Payment

        payments = Payment.objects.filter(
            payment_date__range=[self.start_date, self.end_date]
        ).select_related(
            'job__customer', 'job__worker',
            'booking__customer', 'booking__worker',
        ).order_by('-payment_date')

        payment_rows = []
        for payment in payments:
            customer = payment.job.customer if payment.job else payment.booking.customer if payment.booking else None
            worker = payment.job.worker if payment.job else payment.booking.worker if payment.booking else None
            payment_rows.append([
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.transaction_id or 'N/A',
                customer.get_full_name() if customer else 'N/A',
                f'BDT {payment.customer_amount}',
                f'BDT {payment.platform_commission}',
                payment.payment_status,
                worker.get_full_name() if worker else 'Unassigned',
            ])

        workers = CustomUser.objects.filter(
            role='worker'
        ).select_related('worker_profile').annotate(
            jobs_completed=Count('jobs_as_worker', filter=Q(jobs_as_worker__status='COMPLETED')),
        ).order_by('-worker_profile__total_earnings')
        worker_rows = []
        for worker in workers:
            profile = worker.worker_profile
            worker_rows.append([
                worker.get_full_name() or worker.email,
                profile.verification_status,
                worker.jobs_completed or 0,
                f'BDT {profile.total_earnings}',
                f'BDT {profile.available_earnings}',
            ])

        jobs = Job.objects.filter(
            created_at__range=[self.start_date, self.end_date]
        ).select_related('customer', 'worker').order_by('-created_at')
        job_rows = []
        for job in jobs:
            job_rows.append([
                job.pk,
                job.title,
                job.status,
                job.customer.get_full_name() or job.customer.email,
                job.worker.get_full_name() if job.worker else 'Unassigned',
                f'BDT {job.proposed_price}',
                job.created_at.strftime('%Y-%m-%d'),
            ])

        verified_totals = payments.filter(payment_status='Verified').aggregate(
            revenue=Sum('customer_amount'),
            commission=Sum('platform_commission'),
            worker_earnings=Sum('worker_amount'),
            transactions=Count('pk'),
        )
        financial_rows = [
            ['Total verified revenue', f"BDT {verified_totals['revenue'] or 0}"],
            ['Platform commission', f"BDT {verified_totals['commission'] or 0}"],
            ['Worker earnings', f"BDT {verified_totals['worker_earnings'] or 0}"],
            ['Verified transactions', verified_totals['transactions'] or 0],
        ]

        sections = [
            {
                'title': 'Financial Summary',
                'headers': ['Metric', 'Value'],
                'rows': financial_rows,
            },
            {
                'title': 'Payments',
                'headers': ['Date', 'Transaction', 'Customer', 'Amount', 'Commission', 'Status', 'Worker'],
                'rows': payment_rows or [['No payment records found'] + [''] * 6],
            },
            {
                'title': 'Workers',
                'headers': ['Worker', 'Verification', 'Completed Jobs', 'Total Earnings', 'Available Earnings'],
                'rows': worker_rows or [['No workers found'] + [''] * 4],
            },
            {
                'title': 'Jobs',
                'headers': ['ID', 'Title', 'Status', 'Customer', 'Worker', 'Price', 'Created'],
                'rows': job_rows or [['No jobs found'] + [''] * 6],
            },
        ]

        filename = f"All_Reports_{self.end_date.strftime('%Y%m%d')}.pdf"
        return self.generate_sections_pdf_response(sections, filename)
    
    def generate_financial_summary(self, format_type: str = 'excel') -> HttpResponse:
        """Generate comprehensive financial summary."""
        from payments.models import Payment
        
        verified_payments = Payment.objects.filter(
            payment_date__range=[self.start_date, self.end_date],
            payment_status='Verified'
        )
        
        totals = verified_payments.aggregate(
            total_revenue=Sum('customer_amount'),
            total_commission=Sum('platform_commission'),
            total_worker_earnings=Sum('worker_amount'),
            transaction_count=Count('pk')
        )
        
        headers = ['Metric', 'Value']
        rows = [
            ['Total Revenue', f"{self.CURRENCY}{totals['total_revenue'] or 0}"],
            ['Platform Commission (10%)', f"{self.CURRENCY}{totals['total_commission'] or 0}"],
            ['Worker Earnings (90%)', f"{self.CURRENCY}{totals['total_worker_earnings'] or 0}"],
            ['Number of Transactions', totals['transaction_count'] or 0],
            ['Average Transaction Value', f"{self.CURRENCY}{(totals['total_revenue'] or 0) / (totals['transaction_count'] or 1)}"],
            ['Period', f"{self.start_date.date()} to {self.end_date.date()}"],
        ]
        
        filename = f"Financial_Summary_{self.end_date.strftime('%Y%m%d')}.{format_type}"
        
        if format_type == 'csv':
            return self.generate_csv_response(rows, headers, filename)
        elif format_type == 'excel':
            return self.generate_excel_response(rows, headers, filename)
        else:
            raise ValueError(f"Format {format_type} not supported")
